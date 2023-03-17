from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import openpyxl
import sqlite3
import uuid

def create_pdf_client(company, fio, doc):
    try:
        # Создание временного буфера для pdf документа
        buffer = io.BytesIO()
        pdfmetrics.registerFont(TTFont('ArialUnicode', 'ARIALUNI.ttf'))
        # Создание canvas объекта для формирования pdf документа
        c = canvas.Canvas(buffer, pagesize=letter)
        c.setFont('ArialUnicode', 12) # установка шрифта
        # Добавление текста на pdf страницу
        c.drawString(72, 720, "Заявка на пропуск для клиента")
        c.drawString(72, 700, f"Название компании: {company}")
        c.drawString(72, 680, f"ФИО: {fio}")
        c.drawString(72, 660, f"Номер документа: {doc}")
        # Сохранение pdf документа в буфере
        c.save()
        # Получение байтового массива из буфера
        pdf_data = buffer.getvalue()
        return pdf_data
    except:
        pass

# Функция для создания pdf документа
def create_pdf_loading(company, fio, doc, car_num, tel, fio_man, transit, attorney):
    try:
        # Создание временного буфера для pdf документа
        buffer = io.BytesIO()
        pdfmetrics.registerFont(TTFont('ArialUnicode', 'ARIALUNI.ttf'))
        # Создание canvas объекта для формирования pdf документа
        c = canvas.Canvas(buffer, pagesize=letter)
        c.setFont('ArialUnicode', 12) # установка шрифта
        # Добавление текста на pdf страницу
        c.drawString(72, 720, "Заявка на пропуск для Машины погрузки")
        c.drawString(72, 700, f"Название компании: {company}")
        c.drawString(72, 680, f"ФИО: {fio}")
        c.drawString(72, 660, f"Номер документа: {doc}")
        c.drawString(72, 640, f"Гос номер автомобиля: {car_num}")
        c.drawString(72, 620, f"Контактный телефон: {tel}")
        c.drawString(72, 600, f"ФИО менеджера: {fio_man}")
        c.drawString(72, 580, f"Транзит: {transit}")
        c.drawString(72, 560, f"Доверенность: {attorney}")
        # Сохранение pdf документа в буфере
        c.save()
        # Получение байтового массива из буфера
        pdf_data = buffer.getvalue()
        return pdf_data
    except:
        pass

# Функция для создания pdf документа
def create_pdf_uploading(company, fio, doc, car_num, fio_man, transit, attorney):
    try:
        # Создание временного буфера для pdf документа
        buffer = io.BytesIO()
        pdfmetrics.registerFont(TTFont('ArialUnicode', 'ARIALUNI.ttf'))
        # Создание canvas объекта для формирования pdf документа
        c = canvas.Canvas(buffer, pagesize=letter)
        c.setFont('ArialUnicode', 12) # установка шрифта
        # Добавление текста на pdf страницу
        c.drawString(72, 720, "Заявка на пропуск для Машины выгрузки")
        c.drawString(72, 700, f"Название компании: {company}")
        c.drawString(72, 680, f"ФИО: {fio}")
        c.drawString(72, 660, f"Номер документа: {doc}")
        c.drawString(72, 640, f"Гос номер автомобиля: {car_num}")
        c.drawString(72, 620, f"Контактное лицо: {fio_man}")
        c.drawString(72, 600, f"Транзит: {transit}")
        c.drawString(72, 580, f"Доверенность: {attorney}")
        # Сохранение pdf документа в буфере
        c.save()
        # Получение байтового массива из буфера
        pdf_data = buffer.getvalue()
        return pdf_data
    except:
        pass


# Функция для отправки email
def send_email(pdf_data, theme):
    try:
        # Имя и email администратора
        from_email = 'povarovdd@yandex.ru'
        to_email = 'Admin@keininvest.ru'

        # Создание сообщения
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = theme

        # Добавление PDF файла в сообщение
        pdf_part = MIMEApplication(pdf_data, Name="data.pdf")
        pdf_part['Content-Disposition'] = 'attachment; filename="%s"' % "data.pdf"
        msg.attach(pdf_part)

        with open('pass.docx', 'rb') as f:
            pdf_data = f.read()
            pdf_part = MIMEApplication(pdf_data, Name="pass.docx")
            pdf_part['Content-Disposition'] = 'attachment; filename="%s"' % "pass.docx"
            msg.attach(pdf_part)

        # Отправка email
        smtp_server = "smtp.gmail.com"
        smtp_port = 587
        smtp_username = "uf.noreply.bot@gmail.com"
        smtp_password = "vvldqiwvwsceilhr"

        smtp_server = smtplib.SMTP(smtp_server, smtp_port)
        smtp_server.ehlo() 
        smtp_server.starttls()
        smtp_server.login(smtp_username, smtp_password)
        smtp_server.sendmail("uf.noreply.bot@gmail.com", "Admin@keininvest.ru", msg.as_string())
        smtp_server.quit()
    except:
        pass




def write_data_to_excel_client(company, fio, doc):
    try:
        workbook = openpyxl.load_workbook('work_pass_data.xlsx')

        client_sheet = workbook['client']

        client_sheet.append([company, fio, doc])

        workbook.save('work_pass_data.xlsx')
    except:
        pass


def write_data_to_excel_loading(fio, company, doc, car_num, tel, fio_man, transit, attorney):
    try:
        # Открываем файл Excel
        workbook = openpyxl.load_workbook('work_pass_data.xlsx')

        loading_sheet = workbook['loading']
        # unloading_sheet = workbook['unloading']

        # Записываем данные о клиенте
        loading_sheet.append([fio, company, doc, car_num, tel, fio_man, transit, attorney])

        workbook.save('work_pass_data.xlsx')
    except:
        pass


def write_data_to_excel_uploading(company, fio, doc, car_num, fio_man, transit, attorney):
    try:
        # Открываем файл Excel
        workbook = openpyxl.load_workbook('work_pass_data.xlsx')

        # Получаем нужные листы
        client_sheet = workbook['uploading']

        # Записываем данные о клиенте
        client_sheet.append([fio, company, doc, car_num, fio_man, transit, attorney])

        # Сохраняем файл Excel
        workbook.save('work_pass_data.xlsx')
    except:
        pass

def insert_db_client(company, fio, doc):
    conn = sqlite3.connect('database_uf.db')
    cursor = conn.cursor()
    id = str(uuid.uuid4())
    cursor.execute("INSERT INTO client (company, fio, doc, status, id) VALUES (?, ?, ?, ?, ?)",
                    (company, fio, doc, "0", id))
    conn.commit()
    conn.close()

def insert_db_loading(fio, company, doc, car_num, tel, fio_man, transit, attorney):
    conn = sqlite3.connect('database_uf.db')
    cursor = conn.cursor()
    id = str(uuid.uuid4())
    cursor.execute("INSERT INTO loading (fio, company, doc, car_num, tel, fio_man, transit, attorney, status, id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (fio, company, doc, car_num, tel, fio_man, transit, attorney, "0", id))
    conn.commit()
    conn.close()

def insert_db_uploading(company, fio, doc, car_num, fio_man, transit, attorney):
    conn = sqlite3.connect('database_uf.db')
    cursor = conn.cursor()
    id = str(uuid.uuid4())
    cursor.execute("INSERT INTO uploading (company, fio, doc, car_num, fio_man, transit, attorney, status, id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (company, fio, doc, car_num, fio_man, transit, attorney, "0", id))
    conn.commit()
    conn.close()

